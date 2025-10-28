import csv
from django.http import HttpResponse
from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook

from students.utils import generate_monthly_summary_for_all
from students.models import Complaint, DailyMealStatus, MonthlyMealSummary, Student, StudentMealPreference
from .models import ManagerProfile, MealToken, WeeklyMenuProposal
from students.models import MonthlyMealSummary, WeeklyMenu, WEEKDAY_CHOICES
from .forms import WeeklyMenuProposalForm
from accounts.decorators import manager_required, admin_required
from datetime import date, datetime, timedelta


from .models import SpecialMealRequest
from .forms import SpecialMealRequestForm

def is_manager(user):
    return user.role == "manager"


def is_admin(user):
    return user.role == "admin"


def is_admin_or_manager(user):
    return user.is_authenticated and (
        user.role in ["admin", "manager"] or user.is_superuser
    )


@login_required
@user_passes_test(is_manager)
def manager_menu_list(request):
    proposals = WeeklyMenuProposal.objects.filter(created_by=request.user)
    return render(request, "managers/manager_menu_list.html", {"proposals": proposals})


@manager_required
def propose_weekly_menu(request):
    current_week_start = WeeklyMenuProposal.get_week_start()

    if request.method == "POST":
        # Remove old pending proposals for this week
        WeeklyMenuProposal.objects.filter(
            week_start_date=current_week_start,
            created_by=request.user,
            status="pending",
        ).delete()

        forms_valid = True

        # Loop through 7 days
        for day, _ in WEEKDAY_CHOICES:
            form = WeeklyMenuProposalForm(
                {
                    "day_of_week": day,
                    "week_start_date": current_week_start,
                    "breakfast_main": request.POST.get(f"{day}_breakfast_main", ""),
                    "breakfast_cost": request.POST.get(f"{day}_breakfast_cost", 0),
                    "lunch_main": request.POST.get(f"{day}_lunch_main", ""),
                    "lunch_cost": request.POST.get(f"{day}_lunch_cost", 0),
                    "lunch_contains_beef": request.POST.get(
                        f"{day}_lunch_contains_beef"
                    )
                    == "on",
                    "lunch_contains_fish": request.POST.get(
                        f"{day}_lunch_contains_fish"
                    )
                    == "on",
                    "lunch_alternate": request.POST.get(f"{day}_lunch_alternate", ""),
                    "lunch_cost_alternate": request.POST.get(
                        f"{day}_lunch_cost_alternate", 0
                    ),
                    "dinner_main": request.POST.get(f"{day}_dinner_main", ""),
                    "dinner_cost": request.POST.get(f"{day}_dinner_cost", 0),
                    "dinner_contains_beef": request.POST.get(
                        f"{day}_dinner_contains_beef"
                    )
                    == "on",
                    "dinner_contains_fish": request.POST.get(
                        f"{day}_dinner_contains_fish"
                    )
                    == "on",
                    "dinner_alternate": request.POST.get(f"{day}_dinner_alternate", ""),
                    "dinner_cost_alternate": request.POST.get(
                        f"{day}_dinner_cost_alternate", 0
                    ),
                }
            )

            if form.is_valid():
                proposal = form.save(commit=False)
                proposal.created_by = request.user
                proposal.week_start_date = current_week_start
                proposal.save()
            else:
                forms_valid = False
                print(form.errors)

        # Final message
        if forms_valid:
            messages.success(
                request,
                f"✅ Weekly proposal for week starting {current_week_start} submitted!",
            )
            return redirect("managers:propose_weekly_menu")
        else:
            messages.error(request, "❌ Some fields were invalid. Please check again.")

    return render(
        request,
        "managers/propose_weekly_menu.html",
        {"weekdays": WEEKDAY_CHOICES, "week_start_date": current_week_start},
    )


from django.shortcuts import render
from django.db.models import Sum
from students.models import MonthlyMealSummary, StudentMealPreference
from accounts.decorators import manager_required  # if using custom decorator


@manager_required
def monthly_summary_view(request):
    months = (
        MonthlyMealSummary.objects.values_list("month", flat=True)
        .distinct()
        .order_by("-month")
    )
    selected_month = request.GET.get("month")
    filter_type = request.GET.get("type")  # 'beef_fish' or 'mutton_egg'

    summaries = []
    total_cost = 0

    # preference stats
    beef_count = no_beef_count = fish_count = no_fish_count = 0

    if selected_month:
        # Load summaries for the selected month
        summaries = MonthlyMealSummary.objects.filter(
            month=selected_month
        ).select_related("student")

        # Attach meal preferences for each student
        prefs = {
            p.student_id: p
            for p in StudentMealPreference.objects.filter(month=selected_month)
        }

        for s in summaries:
            pref = prefs.get(s.student.id)
            s.prefers_beef = pref.prefers_beef if pref else True
            s.prefers_fish = pref.prefers_fish if pref else True

        # Compute statistics
        beef_count = sum(1 for s in summaries if s.prefers_beef)
        no_beef_count = sum(1 for s in summaries if not s.prefers_beef)
        fish_count = sum(1 for s in summaries if s.prefers_fish)
        no_fish_count = sum(1 for s in summaries if not s.prefers_fish)

        # Optional filtering for table view
        if filter_type == "beef_fish":
            summaries = [s for s in summaries if s.prefers_beef and s.prefers_fish]
        elif filter_type == "mutton_egg":
            summaries = [
                s for s in summaries if not s.prefers_beef or not s.prefers_fish
            ]

        # Calculate total cost
        total_cost = sum(float(s.total_cost) for s in summaries)

    return render(
        request,
        "managers/monthly_summary.html",
        {
            "months": months,
            "selected_month": selected_month,
            "summaries": summaries,
            "total_cost": total_cost,
            "filter_type": filter_type,
            "beef_count": beef_count,
            "no_beef_count": no_beef_count,
            "fish_count": fish_count,
            "no_fish_count": no_fish_count,
        },
    )


def _get_summaries_and_prefs(month):
    summaries = (
        MonthlyMealSummary.objects.filter(month=month)
        .select_related("student")
        .order_by("student__room_number")
    )
    prefs = {p.student_id: p for p in StudentMealPreference.objects.filter(month=month)}
    return summaries, prefs


def _determine_pref_values(s, prefs):
    pref = prefs.get(s.student.id)
    prefers_beef = pref.prefers_beef if pref else getattr(s.student, "default_prefers_beef", True)
    prefers_fish = pref.prefers_fish if pref else getattr(s.student, "default_prefers_fish", True)
    return prefers_beef, prefers_fish


def _meal_type_label(prefers_beef, prefers_fish):
    if prefers_beef and prefers_fish:
        return "Beef + Fish"
    if not prefers_beef and prefers_fish:
        return "Mutton + Fish"
    if prefers_beef and not prefers_fish:
        return "Beef + Egg"
    return "Mutton + Egg"


def _build_workbook(summaries, prefs, month):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Summary_{month}"

    headers = [
        "Room Number",
        "Student Name",
        "Total ON Days",
        "Prefers Beef",
        "Prefers Fish",
        "Meal Type",
        "Total Cost (৳)",
    ]
    ws.append(headers)

    for s in summaries:
        prefers_beef, prefers_fish = _determine_pref_values(s, prefs)
        meal_type = _meal_type_label(prefers_beef, prefers_fish)

        ws.append(
            [
                s.student.room_number,
                s.student.name,
                s.total_on_days,
                "Yes" if prefers_beef else "No",
                "Yes" if prefers_fish else "No",
                meal_type,
                float(s.total_cost),
            ]
        )

    # Style header row
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    return wb


@manager_required
def export_monthly_summary(request):
    month = request.GET.get("month")
    if not month:
        return HttpResponse("Month parameter missing.", status=400)

    summaries, prefs = _get_summaries_and_prefs(month)
    wb = _build_workbook(summaries, prefs, month)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Meal_Summary_{month}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@manager_required
def regenerate_monthly_summary(request, month):
    from students.utils import generate_monthly_summary_for_all

    y, m = map(int, month.split("-"))
    generate_monthly_summary_for_all(y, m)
    messages.success(request, f"Monthly summary regenerated for {month}")
    return redirect("managers:monthly_summary")


@manager_required
def manage_complaints(request):
    complaints = Complaint.objects.all().order_by("-created_at")

    if request.method == "POST":
        complaint_id = request.POST.get("complaint_id")
        complaint = Complaint.objects.get(id=complaint_id)
        complaint.is_fixed = not complaint.is_fixed
        complaint.fixed_by = request.user if complaint.is_fixed else None
        complaint.fixed_at = timezone.now() if complaint.is_fixed else None
        complaint.save()
        messages.success(request, "Complaint status updated successfully!")
        return redirect("managers:manage_complaints")

    return render(
        request, "managers/manage_complaints.html", {"complaints": complaints}
    )


@admin_required
def create_special_request(request):
    if request.method == "POST":
        form = SpecialMealRequestForm(request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req.created_by = request.user
            req.save()
            messages.success(request, "Special meal request submitted to the manager.")
            return redirect("managers:create_special_request")
    else:
        form = SpecialMealRequestForm()
    return render(
        request,
        "managers/create_special_request.html",
        {"form": form, "page_title": "Create Special Meal Request"},
    )


@user_passes_test(is_admin_or_manager)
def manager_requests_list(request):
    if request.user.role == "admin":
        # Admin can see all requests
        requests_qs = SpecialMealRequest.objects.all().order_by("-created_at")
    else:
        requests_qs = SpecialMealRequest.objects.filter(manager=request.user).order_by("-created_at")

    return render(
        request,
        "managers/manager_requests_list.html",
        {"requests": requests_qs, "page_title": "Special Meal Requests"},
    )


@manager_required
def manager_request_detail(request, pk):
    req = get_object_or_404(SpecialMealRequest, pk=pk, manager=request.user)

    if request.method == "POST":
        note = request.POST.get("response_note", "")

        # marking as accepted
        req.mark_responded(SpecialMealRequest.STATUS_ACCEPTED, note)
        messages.success(request, "Special meal request marked as accepted.")
        return redirect("managers:manager_requests_list")

    return render(
        request,
        "managers/manager_request_detail.html",
        {"request_obj": req, "page_title": "Request Detail"},
    )


@login_required
@user_passes_test(is_manager)
def manager_profile_view(request):
    """Display the logged-in manager's full profile info"""
    manager = get_object_or_404(ManagerProfile, user=request.user)

    context = {
        "page_title": "My Profile",
        "manager": manager,
    }
    return render(request, "managers/profile.html", context)


@login_required
@user_passes_test(manager_required)
def search_students_by_room(request):
    students = []
    room_number = request.GET.get("room_number")
    today = timezone.localdate()

    if room_number:
        students = Student.objects.filter(room_number=room_number)
        for student in students:
            # Today's status
            student.today_status = DailyMealStatus.objects.filter(
                student=student, date=today
            ).first()

            # Current month's preference
            student.current_pref = StudentMealPreference.objects.filter(
                student=student, month=today.strftime("%Y-%m")
            ).first()

            # Tokens already issued today
            tokens = MealToken.objects.filter(student=student, date=today)
            student.issued_meals = {token.meal_type for token in tokens}

    meals = ["breakfast", "lunch", "dinner"]

    context = {
        "students": students,
        "room_number": room_number,
        "today": today,
        "meals": meals,
    }
    return render(request, "managers/search_students.html", context)


@login_required
@user_passes_test(manager_required)
def issue_token(request, student_id, meal_type):
    student = get_object_or_404(Student, id=student_id)
    today = timezone.localdate()

    # Check if token already exists
    if MealToken.objects.filter(
        student=student, date=today, meal_type=meal_type
    ).exists():
        messages.warning(
            request, f"Token already issued for {student.name} ({meal_type})"
        )
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Get student meal preference for current month
    pref = StudentMealPreference.objects.filter(
        student=student, month=today.strftime("%Y-%m")
    ).first()

    # Get today's menu
    weekly_menu = WeeklyMenu.objects.get(day_of_week=today.strftime("%A"))

    # Determine token type
    token_type = "main"
    if (
        (
            meal_type == "lunch"
            and weekly_menu.lunch_contains_beef
            and not pref.prefers_beef
        )
        or (
            meal_type == "lunch"
            and weekly_menu.lunch_contains_fish
            and not pref.prefers_fish
        )
        or (
            meal_type == "dinner"
            and weekly_menu.dinner_contains_beef
            and not pref.prefers_beef
        )
        or (
            meal_type == "dinner"
            and weekly_menu.dinner_contains_fish
            and not pref.prefers_fish
        )
    ):
        token_type = "alternate"

    # Create token
    MealToken.objects.create(
        student=student,
        date=today,
        meal_type=meal_type,
        token_type=token_type,
        issued_by=request.user,
        collected=True,
    )

    messages.success(
        request, f"{token_type.title()} token issued for {student.name} ({meal_type})"
    )
    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
@user_passes_test(manager_required)
def daily_token_summary(request):
    today = timezone.localdate()
    tokens = MealToken.objects.filter(date=today)

    summary = {
        "breakfast_main": tokens.filter(
            meal_type="breakfast", token_type="main"
        ).count(),
        "breakfast_alt": tokens.filter(
            meal_type="breakfast", token_type="alternate"
        ).count(),
        "lunch_main": tokens.filter(meal_type="lunch", token_type="main").count(),
        "lunch_alt": tokens.filter(meal_type="lunch", token_type="alternate").count(),
        "dinner_main": tokens.filter(meal_type="dinner", token_type="main").count(),
        "dinner_alt": tokens.filter(meal_type="dinner", token_type="alternate").count(),
        "total_tokens": tokens.count(),
    }

    return render(
        request,
        "managers/daily_token_summary.html",
        {
            "summary": summary,
            "tokens": tokens,
            "today": today,
            "page_title": "Daily Token Summary",
        },
    )


@login_required
@user_passes_test(manager_required)
def export_daily_token_summary(request):
    # Get date from query params or use today
    date_str = request.GET.get("date")
    if date_str:
        from datetime import datetime

        try:
            date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            date = timezone.localdate()
    else:
        date = timezone.localdate()

    tokens = MealToken.objects.filter(date=date)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tokens_{date}"

    # Header row
    headers = ["Student", "Room", "Meal", "Token Type", "Issued At", "Issued By"]
    ws.append(headers)

    # Data rows
    for token in tokens:
        ws.append(
            [
                token.student.name,
                token.student.room_number,
                token.meal_type.capitalize(),
                token.token_type.capitalize(),
                token.issued_at.strftime("%Y-%m-%d %H:%M:%S"),
                token.issued_by.username if token.issued_by else "N/A",
            ]
        )

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="daily_token_summary_{date}.xlsx"'
    )
    wb.save(response)
    return response
